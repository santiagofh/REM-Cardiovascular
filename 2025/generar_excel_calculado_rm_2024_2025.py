from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path(
    r"C:\Users\fariass\OneDrive - SUBSECRETARIA DE SALUD PUBLICA\Escritorio\REM\REM-Cardiovascular"
)
DATA_DIR = ROOT / "2025"
TEMPLATE_PATH = ROOT / "Planilla indicadores y fechas reuniones macrozonales 2026.xlsx"
RM_PATH = DATA_DIR / "indicadores_cardiovascular_rm_2024_2025.csv"
EGRESOS_PATH = DATA_DIR / "egresos_hospitalarios_factibilidad_resumen_2020_2024.csv"
OUTPUT_PATH = ROOT / "Planilla indicadores calculados RM 2024-2025.xlsx"

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")


def format_value(value: object) -> object:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return None
    return float(numeric) / 100.0


def format_count(value: object) -> object:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return None
    return float(numeric)


def normalize_indicator_id(indicator_id: object, indicator_name: object) -> str:
    text = str(indicator_id).strip()
    name = str(indicator_name or "").lower()
    if text == "12":
        if "fondo de ojo" in name:
            return "12b"
        if "tamizaje" in name:
            return "12a"
    return text


def copy_style(ws, source_col: int, target_col: int, max_row: int) -> None:
    for row in range(1, max_row + 1):
        source = ws.cell(row, source_col)
        target = ws.cell(row, target_col)
        if source.has_style:
            target._style = source._style
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = source.font.copy()
        if source.border:
            target.border = source.border.copy()
        if source.alignment:
            target.alignment = source.alignment.copy()
        if source.fill:
            target.fill = source.fill.copy()


def build_rm_lookup() -> dict[tuple[str, int], dict[str, object]]:
    rm = pd.read_csv(RM_PATH, dtype={"indicador_id": str})
    lookup: dict[tuple[str, int], dict[str, object]] = {}
    for _, row in rm.iterrows():
        key = (str(row["indicador_id"]).strip(), int(row["Ano"]))
        lookup[key] = {
            "valor": format_value(row["valor"]),
            "numerador": format_count(row["numerador"]),
            "denominador": format_count(row["denominador"]),
            "estado": str(row["estado_calculo"]).strip(),
        }
    return lookup


def build_egresos_lookup() -> dict[tuple[str, int], dict[str, object]]:
    if not EGRESOS_PATH.exists():
        return {}
    eg = pd.read_csv(EGRESOS_PATH, dtype={"indicador": str})
    lookup: dict[tuple[str, int], dict[str, object]] = {}
    for _, row in eg.iterrows():
        indicator = str(row["indicador"]).strip().replace("_proxy", "")
        if indicator not in {"18", "19", "20", "21", "22"}:
            continue
        year = int(row["ano"])
        lookup[(indicator, year)] = {
            "conteo": format_count(row["total"]),
            "estado": str(row["estado"]).strip(),
        }
    return lookup


def main() -> None:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Indicadores"]

    # Add calculated columns after existing N/D reference columns.
    ws.insert_cols(26, amount=6)
    copy_style(ws, 24, 26, ws.max_row)
    copy_style(ws, 25, 27, ws.max_row)
    copy_style(ws, 24, 28, ws.max_row)
    copy_style(ws, 25, 29, ws.max_row)
    copy_style(ws, 24, 30, ws.max_row)
    copy_style(ws, 25, 31, ws.max_row)

    headers = {
        23: "Valor 2024 calc",
        24: "N 2024",
        25: "D 2024",
        26: "Valor 2025 calc",
        27: "N 2025",
        28: "D 2025",
        29: "Estado 2024",
        30: "Estado 2025",
        31: "Nota",
    }
    for col, label in headers.items():
        ws.cell(1, col).value = label
        ws.cell(1, col).fill = YELLOW_FILL

    # Blank old historical series so only calculated output remains visible.
    for row in range(2, ws.max_row + 1):
        for col in range(7, 13):
            ws.cell(row, col).value = None

    rm_lookup = build_rm_lookup()
    egresos_lookup = build_egresos_lookup()

    for row in range(2, ws.max_row + 1):
        raw_indicator = ws.cell(row, 1).value
        indicator_name = ws.cell(row, 2).value
        indicator_id = normalize_indicator_id(raw_indicator, indicator_name)

        v2024 = rm_lookup.get((indicator_id, 2024))
        v2025 = rm_lookup.get((indicator_id, 2025))

        # Use main year columns of the template for calculated values.
        ws.cell(row, 11).value = None if v2024 is None else v2024["valor"]
        ws.cell(row, 12).value = None if v2025 is None else v2025["valor"]
        ws.cell(row, 11).number_format = "0.0%" if v2024 is not None else ws.cell(row, 11).number_format
        ws.cell(row, 12).number_format = "0.0%" if v2025 is not None else ws.cell(row, 12).number_format

        ws.cell(row, 23).value = None if v2024 is None else v2024["valor"]
        ws.cell(row, 24).value = None if v2024 is None else v2024["numerador"]
        ws.cell(row, 25).value = None if v2024 is None else v2024["denominador"]
        ws.cell(row, 26).value = None if v2025 is None else v2025["valor"]
        ws.cell(row, 27).value = None if v2025 is None else v2025["numerador"]
        ws.cell(row, 28).value = None if v2025 is None else v2025["denominador"]
        ws.cell(row, 29).value = "" if v2024 is None else str(v2024["estado"]).capitalize()
        ws.cell(row, 30).value = "" if v2025 is None else str(v2025["estado"]).capitalize()

        for col in [23, 26]:
            ws.cell(row, col).number_format = "0.0%"
            ws.cell(row, col).fill = GREEN_FILL
        for col in [24, 25, 27, 28, 29, 30, 31]:
            ws.cell(row, col).fill = GREEN_FILL

        note = ""
        if indicator_id == "12a":
            note = "No calculado localmente: la planilla usa una logica de tamizaje RD distinta a fondo de ojo."
        elif indicator_id in {"18", "19", "20", "21"}:
            count_2024 = egresos_lookup.get((indicator_id, 2024))
            if count_2024 is not None:
                ws.cell(row, 24).value = count_2024["conteo"]
                ws.cell(row, 29).value = "Conteo 2024"
                note = (
                    "El archivo local permite contar egresos FONASA 15+ RM 2024, "
                    "pero no calcular la tasa exacta porque aqui no esta el denominador oficial de beneficiarios FONASA 15+."
                )
            else:
                note = "Sin base local suficiente para tasa."
        elif indicator_id == "22":
            note = (
                "No calculado: el CSV local de egresos no trae procedimiento/intervencion para identificar amputacion de pie diabetico."
            )
        elif indicator_id in {"16", "17"} and v2024 is not None and str(v2024["estado"]).lower() == "proxy":
            note = "2024 usa proxy con codigo consolidado historico."

        ws.cell(row, 31).value = note

    ws["K1"] = 2024
    ws["L1"] = 2025
    ws.freeze_panes = "B2"

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
