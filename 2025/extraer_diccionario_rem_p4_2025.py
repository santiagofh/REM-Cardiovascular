from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PRODUCTO_ROOT = Path(
    r"C:\Users\fariass\OneDrive - SUBSECRETARIA DE SALUD PUBLICA\Escritorio\REM\REM-Cardiovascular"
)
ANIO = "2025"
REM = "P4"
DICCIONARIO_DIR = PRODUCTO_ROOT / ANIO
DICT_PATH = Path(
    r"C:\Users\fariass\OneDrive - SUBSECRETARIA DE SALUD PUBLICA\Escritorio\DATA\REM\REM_2025\Diccionarios\DICCIONARIO CODIGOS SP_25_V1.2.xlsm"
)
INDICADORES_PATH = (
    PRODUCTO_ROOT / "Planilla indicadores y fechas reuniones macrozonales 2026.xlsx"
)
CSV_ENCODING = "utf-8-sig"


INDICADORES_REM_P4 = [
    {
        "indicador": 1,
        "nombre": "Cobertura de hipertensión arterial (HTA)",
        "rol": "numerador",
        "codigo_prestacion": "P4150601",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Hipertensión arterial.",
    },
    {
        "indicador": 2,
        "nombre": "Control de HTA",
        "rol": "numerador",
        "codigo_prestacion": "P4180200",
        "nota_metodologica": "Sección B -> Personas bajo control por hipertensión -> PA < 140/90 mmHg.",
    },
    {
        "indicador": 2,
        "nombre": "Control de HTA",
        "rol": "numerador",
        "codigo_prestacion": "P4200100",
        "nota_metodologica": "Sección B -> Personas bajo control por hipertensión -> PA < 150/90 mmHg.",
    },
    {
        "indicador": 2,
        "nombre": "Control de HTA",
        "rol": "denominador",
        "codigo_prestacion": "P4150601",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Hipertensión arterial.",
    },
    {
        "indicador": 3,
        "nombre": "Cobertura efectiva (tasa de control poblacional) de HTA",
        "rol": "numerador",
        "codigo_prestacion": "P4180200",
        "nota_metodologica": "Sección B -> Personas bajo control por hipertensión -> PA < 140/90 mmHg.",
    },
    {
        "indicador": 3,
        "nombre": "Cobertura efectiva (tasa de control poblacional) de HTA",
        "rol": "numerador",
        "codigo_prestacion": "P4200100",
        "nota_metodologica": "Sección B -> Personas bajo control por hipertensión -> PA < 150/90 mmHg.",
    },
    {
        "indicador": 4,
        "nombre": "Porcentaje de personas con diagnóstico de HTA, muy descompensadas",
        "rol": "numerador",
        "codigo_prestacion": "P4200400",
        "nota_metodologica": "Sección C -> Personas con hipertensión en PSCV -> Con presión arterial igual o mayor a 160/100 mmHg.",
    },
    {
        "indicador": 4,
        "nombre": "Porcentaje de personas con diagnóstico de HTA, muy descompensadas",
        "rol": "denominador",
        "codigo_prestacion": "P4150601",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Hipertensión arterial.",
    },
    {
        "indicador": 6,
        "nombre": "Cobertura de diabetes mellitus tipo 2 (DM2)",
        "rol": "numerador",
        "codigo_prestacion": "P4150602",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Diabetes mellitus tipo 2.",
    },
    {
        "indicador": 7,
        "nombre": "Control de DM2",
        "rol": "numerador",
        "codigo_prestacion": "P4180300",
        "nota_metodologica": "Sección B -> Personas bajo control por diabetes mellitus -> HbA1C < 7%.",
    },
    {
        "indicador": 7,
        "nombre": "Control de DM2",
        "rol": "numerador",
        "codigo_prestacion": "P4200200",
        "nota_metodologica": "Sección B -> Personas bajo control por diabetes mellitus -> HbA1C < 8%.",
    },
    {
        "indicador": 7,
        "nombre": "Control de DM2",
        "rol": "denominador",
        "codigo_prestacion": "P4150602",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Diabetes mellitus tipo 2.",
    },
    {
        "indicador": 8,
        "nombre": "Cobertura efectiva (tasa de control poblacional) de DM2",
        "rol": "numerador",
        "codigo_prestacion": "P4180300",
        "nota_metodologica": "Sección B -> Personas bajo control por diabetes mellitus -> HbA1C < 7%.",
    },
    {
        "indicador": 8,
        "nombre": "Cobertura efectiva (tasa de control poblacional) de DM2",
        "rol": "numerador",
        "codigo_prestacion": "P4200200",
        "nota_metodologica": "Sección B -> Personas bajo control por diabetes mellitus -> HbA1C < 8%.",
    },
    {
        "indicador": 9,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, muy descompensadas",
        "rol": "numerador",
        "codigo_prestacion": "P4190960",
        "nota_metodologica": "Sección C -> Personas con diabetes en PSCV -> Con HBA1C >= 9 %.",
    },
    {
        "indicador": 9,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, muy descompensadas",
        "rol": "denominador",
        "codigo_prestacion": "P4150602",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Diabetes mellitus tipo 2.",
    },
    {
        "indicador": 10,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, compensada, usuarias de insulina",
        "rol": "numerador",
        "codigo_prestacion": "P4200700",
        "nota_metodologica": "Sección C -> Personas con diabetes en PSCV -> En tratamiento con insulina que logra meta con HBA1C según edad.",
    },
    {
        "indicador": 10,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, compensada, usuarias de insulina",
        "rol": "denominador",
        "codigo_prestacion": "P4180800",
        "nota_metodologica": "Sección C -> Personas con diabetes en PSCV -> En tratamiento con insulina.",
    },
    {
        "indicador": 11,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de pie diabético vigente",
        "rol": "numerador",
        "codigo_prestacion": "P4190809",
        "nota_metodologica": "Sección C -> Evaluación del pie diabético vigente -> Riesgo bajo.",
    },
    {
        "indicador": 11,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de pie diabético vigente",
        "rol": "numerador",
        "codigo_prestacion": "P4170300",
        "nota_metodologica": "Sección C -> Evaluación del pie diabético vigente -> Riesgo moderado.",
    },
    {
        "indicador": 11,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de pie diabético vigente",
        "rol": "numerador",
        "codigo_prestacion": "P4190500",
        "nota_metodologica": "Sección C -> Evaluación del pie diabético vigente -> Riesgo alto.",
    },
    {
        "indicador": 11,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de pie diabético vigente",
        "rol": "numerador",
        "codigo_prestacion": "P4190600",
        "nota_metodologica": "Sección C -> Evaluación del pie diabético vigente -> Riesgo máximo.",
    },
    {
        "indicador": 11,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de pie diabético vigente",
        "rol": "denominador",
        "codigo_prestacion": "P4150602",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Diabetes mellitus tipo 2.",
    },
    {
        "indicador": 12,
        "nombre": "Porcentaje de personas con diagnóstico de DM2 y tamizaje de RD vigente",
        "rol": "numerador",
        "codigo_prestacion": "P4190950",
        "nota_metodologica": "La planilla menciona tamizaje de RD; en el diccionario P4 vigente la prestación observable es 'Con fondo de ojo, vigente'.",
    },
    {
        "indicador": 12,
        "nombre": "Porcentaje de personas con diagnóstico de DM2 y tamizaje de RD vigente",
        "rol": "denominador",
        "codigo_prestacion": "P4150602",
        "nota_metodologica": "Base de personas con diabetes mellitus tipo 2 bajo control.",
    },
    {
        "indicador": 12,
        "nombre": "Porcentaje de personas con diagnóstico de DM2 y tamizaje de RD vigente",
        "rol": "ajuste_denominador",
        "codigo_prestacion": "P4302102",
        "nota_metodologica": "La planilla descuenta retinopatía diabética en el denominador; este código queda marcado para esa depuración.",
    },
    {
        "indicador": 12,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de fondo de ojo (FO) vigente",
        "rol": "numerador",
        "codigo_prestacion": "P4190950",
        "nota_metodologica": "Sección C -> Personas con diabetes en PSCV -> Con fondo de ojo, vigente.",
    },
    {
        "indicador": 12,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de fondo de ojo (FO) vigente",
        "rol": "denominador",
        "codigo_prestacion": "P4150602",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Diabetes mellitus tipo 2.",
    },
    {
        "indicador": 13,
        "nombre": "Porcentaje de personas con diagnóstico de HTA, con evaluación de función renal",
        "rol": "numerador",
        "codigo_prestacion": "P4301080",
        "nota_metodologica": "Sección C -> Personas con hipertensión en PSCV -> Con VFGE y RAC vigente.",
    },
    {
        "indicador": 13,
        "nombre": "Porcentaje de personas con diagnóstico de HTA, con evaluación de función renal",
        "rol": "denominador",
        "codigo_prestacion": "P4150601",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Hipertensión arterial.",
    },
    {
        "indicador": 14,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de función renal",
        "rol": "numerador",
        "codigo_prestacion": "P4301040",
        "nota_metodologica": "Sección C -> Personas con diabetes en PSCV -> Con VFGE y RAC vigente.",
    },
    {
        "indicador": 14,
        "nombre": "Porcentaje de personas con diagnóstico de DM2, con evaluación de función renal",
        "rol": "denominador",
        "codigo_prestacion": "P4150602",
        "nota_metodologica": "Sección A -> Personas bajo control según patología y factores de riesgo -> Diabetes mellitus tipo 2.",
    },
    {
        "indicador": 15,
        "nombre": "Porcentaje de personas con diagnóstico de DM y ERC en tratamiento de prevención secundaria de ERC",
        "rol": "numerador",
        "codigo_prestacion": "P4401019",
        "nota_metodologica": "Sección C -> Personas con diabetes en PSCV -> Con ERC y en tratamiento con IECA o ARA II.",
    },
    {
        "indicador": 15,
        "nombre": "Porcentaje de personas con diagnóstico de DM y ERC en tratamiento de prevención secundaria de ERC",
        "rol": "denominador",
        "codigo_prestacion": "P4301070",
        "nota_metodologica": "Sección C -> Personas con diabetes en PSCV -> Con diagnóstico de enfermedad renal crónica.",
    },
    {
        "indicador": 16,
        "nombre": "Porcentaje de personas con diagnóstico de enfermedad cardio-cerebrovascular (ECV), en tratamiento con antiagregante plaquetario",
        "rol": "numerador",
        "codigo_prestacion": "P4401013",
        "nota_metodologica": "Sección B -> Personas con antecedentes de IAM -> En tratamiento con antiagregantes plaquetarios.",
    },
    {
        "indicador": 16,
        "nombre": "Porcentaje de personas con diagnóstico de enfermedad cardio-cerebrovascular (ECV), en tratamiento con antiagregante plaquetario",
        "rol": "numerador",
        "codigo_prestacion": "P4401016",
        "nota_metodologica": "Sección B -> Personas con antecedentes de ACV -> En tratamiento con antiagregantes plaquetarios.",
    },
    {
        "indicador": 16,
        "nombre": "Porcentaje de personas con diagnóstico de enfermedad cardio-cerebrovascular (ECV), en tratamiento con antiagregante plaquetario",
        "rol": "denominador",
        "codigo_prestacion": "P4190900",
        "nota_metodologica": "Sección A -> Antecedentes de infarto agudo al miocardio (IAM).",
    },
    {
        "indicador": 16,
        "nombre": "Porcentaje de personas con diagnóstico de enfermedad cardio-cerebrovascular (ECV), en tratamiento con antiagregante plaquetario",
        "rol": "denominador",
        "codigo_prestacion": "P4190910",
        "nota_metodologica": "Sección A -> Antecedentes de ataque cerebrovascular (ACV).",
    },
    {
        "indicador": 17,
        "nombre": "Porcentaje de personas con diagnóstico de ECV, en tratamiento con estatinas",
        "rol": "numerador",
        "codigo_prestacion": "P4401014",
        "nota_metodologica": "Sección B -> Personas con antecedentes de IAM -> En tratamiento con estatina.",
    },
    {
        "indicador": 17,
        "nombre": "Porcentaje de personas con diagnóstico de ECV, en tratamiento con estatinas",
        "rol": "numerador",
        "codigo_prestacion": "P4401017",
        "nota_metodologica": "Sección B -> Personas con antecedentes de ACV -> En tratamiento con estatina.",
    },
    {
        "indicador": 17,
        "nombre": "Porcentaje de personas con diagnóstico de ECV, en tratamiento con estatinas",
        "rol": "denominador",
        "codigo_prestacion": "P4190900",
        "nota_metodologica": "Sección A -> Antecedentes de infarto agudo al miocardio (IAM).",
    },
    {
        "indicador": 17,
        "nombre": "Porcentaje de personas con diagnóstico de ECV, en tratamiento con estatinas",
        "rol": "denominador",
        "codigo_prestacion": "P4190910",
        "nota_metodologica": "Sección A -> Antecedentes de ataque cerebrovascular (ACV).",
    },
]


INDICADORES_NO_REM = [
    {
        "indicador": 5,
        "nombre": "Cobertura de evaluación del Índice de Madurez HEARTS",
        "origen": "No REM",
        "fuente": "Fuente de información respectiva del índice HEARTS.",
    },
    {
        "indicador": 18,
        "nombre": "Tasa de egresos hospitalarios por enfermedad cerebrovascular",
        "origen": "DEIS/FONASA",
        "fuente": "Tablero DEIS Egresos Hospitalarios + población beneficiaria FONASA.",
    },
    {
        "indicador": 19,
        "nombre": "Tasa de egresos hospitalarios por enfermedades isquémicas del corazón",
        "origen": "DEIS/FONASA",
        "fuente": "Tablero DEIS Egresos Hospitalarios + población beneficiaria FONASA.",
    },
    {
        "indicador": 20,
        "nombre": "Tasa de egresos hospitalarios por insuficiencia cardiaca",
        "origen": "DEIS/FONASA",
        "fuente": "Tablero DEIS Egresos Hospitalarios + población beneficiaria FONASA.",
    },
    {
        "indicador": 21,
        "nombre": "Tasa de egresos hospitalarios por diabetes mellitus",
        "origen": "DEIS/FONASA",
        "fuente": "Tablero DEIS Egresos Hospitalarios + población beneficiaria FONASA.",
    },
    {
        "indicador": 22,
        "nombre": "Tasa de egresos hospitalarios en personas con diabetes mellitus y amputación del pie diabético de establecimientos pertenecientes al SNSS",
        "origen": "DEIS/FONASA",
        "fuente": "Tablero PSCV DEIS + población beneficiaria FONASA.",
    },
]


@dataclass(frozen=True)
class SectionTemplate:
    key: str
    name: str
    title: str
    start_row: int
    end_row: int
    first_code_row: int
    last_code_row: int
    row_to_code: dict[int, str]
    code_to_detail: dict[str, str]
    codes: list[str]
    placeholder_to_col: dict[str, int]
    columns_by_placeholder: dict[str, str]


def normalizar_espacios(texto: str) -> str:
    return re.sub(r"\s+", " ", texto).strip()


def normalizar_texto(valor: object) -> str:
    if pd.isna(valor):
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    return texto


def normalizar_codigo_prestacion(valor: object) -> str:
    return normalizar_texto(valor).upper()


def normalizar_texto_visible(valor: object) -> str:
    texto = normalizar_espacios(normalizar_texto(valor))
    if not texto or texto.startswith("="):
        return ""
    return texto


def es_encabezado_seccion(valor: object) -> bool:
    if valor is None:
        return False
    texto = str(valor).upper()
    return "SECCIÓN" in texto or "SECCION" in texto


def token_seccion(texto: str) -> str:
    match = re.search(r"SECCI[ÓO]N\s+([A-Z](?:\.\d+|\d*)?)", texto.upper())
    if not match:
        return "SECCION"
    return match.group(1).replace(" ", "")


def nombre_seccion(texto: str, token: str) -> str:
    limpio = normalizar_espacios(texto)
    if ":" in limpio:
        return limpio.split(":", 1)[1].strip()
    return token


def es_col_placeholder(valor: object) -> bool:
    return isinstance(valor, str) and re.fullmatch(r"COL\d{2}", valor.strip().upper()) is not None


def indice_placeholder(valor: str) -> int:
    return int(valor.strip().upper().replace("COL", ""))


def paths_metadata(rem: str) -> dict[str, Path]:
    prefijo = f"REM_{rem}_{ANIO}"
    return {
        "secciones": DICCIONARIO_DIR / f"{prefijo}_seccion_a_codigos.json",
        "detalle": DICCIONARIO_DIR / f"{prefijo}_codigo_a_detalle.json",
        "columnas_json": DICCIONARIO_DIR / f"{prefijo}_columnas_por_codigo.json",
        "columnas_csv": DICCIONARIO_DIR / f"{prefijo}_columnas_por_codigo.csv",
        "prestaciones_csv": DICCIONARIO_DIR / f"{prefijo}_prestaciones.csv",
        "prestaciones_json": DICCIONARIO_DIR / f"{prefijo}_prestaciones_completo.json",
    }


def build_merged_lookup(ws) -> dict[tuple[int, int], object]:
    lookup: dict[tuple[int, int], object] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = ws.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                lookup[(row, col)] = value
    return lookup


def valor_visible_celda(ws, merged_lookup: dict[tuple[int, int], object], row: int, col: int) -> object:
    value = ws.cell(row, col).value
    if value not in (None, ""):
        return value
    return merged_lookup.get((row, col))


def detectar_secciones(ws, merged_lookup: dict[tuple[int, int], object]) -> list[tuple[int, str]]:
    secciones: list[tuple[int, str]] = []
    for row in range(1, ws.max_row + 1):
        texto = valor_visible_celda(ws, merged_lookup, row, 2)
        if not es_encabezado_seccion(texto):
            continue
        secciones.append((row, normalizar_espacios(str(texto))))
    return secciones


def construir_etiqueta_columna(
    ws,
    merged_lookup: dict[tuple[int, int], object],
    start_row: int,
    first_code_row: int,
    col: int,
) -> str:
    partes: list[str] = []
    for row in range(start_row + 1, first_code_row):
        texto = normalizar_texto_visible(valor_visible_celda(ws, merged_lookup, row, col))
        if not texto or es_col_placeholder(texto) or es_encabezado_seccion(texto):
            continue
        if partes and texto == partes[-1]:
            continue
        partes.append(texto)
    return " - ".join(partes)


def construir_detalle_prestacion(
    ws,
    merged_lookup: dict[tuple[int, int], object],
    row: int,
    first_placeholder_col: int,
    descriptor_context: dict[int, str],
) -> str:
    partes: list[str] = []
    for col in range(2, first_placeholder_col):
        texto = normalizar_texto_visible(valor_visible_celda(ws, merged_lookup, row, col))
        if texto:
            descriptor_context[col] = texto
        texto_contexto = descriptor_context.get(col, "")
        if not texto_contexto:
            continue
        if partes and texto_contexto == partes[-1]:
            continue
        partes.append(texto_contexto)
    return " - ".join(partes)


def extraer_plantilla_rem(rem: str, dict_path: Path, sheet_name: str) -> tuple[list[SectionTemplate], list[dict], list[dict]]:
    wb = load_workbook(dict_path, data_only=False, keep_vba=False)
    ws = wb[sheet_name]
    merged_lookup = build_merged_lookup(ws)
    secciones = detectar_secciones(ws, merged_lookup)

    plantillas: list[SectionTemplate] = []
    prestaciones_rows: list[dict] = []
    columnas_rows: list[dict] = []

    for idx, (start_row, title) in enumerate(secciones):
        hard_end = secciones[idx + 1][0] - 1 if idx + 1 < len(secciones) else ws.max_row
        key = token_seccion(title)
        section_name = nombre_seccion(title, key)

        row_to_code: dict[int, str] = {}
        placeholder_to_col: dict[str, int] = {}
        first_code_row: int | None = None
        last_code_row = start_row

        for row in range(start_row, hard_end + 1):
            code = normalizar_codigo_prestacion(ws.cell(row, 1).value)
            placeholders_en_fila: list[str] = []
            for col in range(2, ws.max_column + 1):
                valor = valor_visible_celda(ws, merged_lookup, row, col)
                if not es_col_placeholder(valor):
                    continue
                placeholder = str(valor).strip().upper()
                placeholders_en_fila.append(placeholder)
                placeholder_to_col.setdefault(placeholder, col)
            if code and placeholders_en_fila:
                row_to_code[row] = code
                if first_code_row is None:
                    first_code_row = row
                last_code_row = row

        if not row_to_code or first_code_row is None:
            continue

        columns_by_placeholder = {
            placeholder: construir_etiqueta_columna(
                ws,
                merged_lookup,
                start_row,
                first_code_row,
                col,
            )
            for placeholder, col in sorted(
                placeholder_to_col.items(),
                key=lambda item: indice_placeholder(item[0]),
            )
        }

        descriptor_context: dict[int, str] = {}
        code_to_detail: dict[str, str] = {}
        codes: list[str] = []

        for row in range(first_code_row, last_code_row + 1):
            code = row_to_code.get(row)
            if not code:
                continue

            placeholder_cols = [
                col
                for placeholder, col in placeholder_to_col.items()
                if es_col_placeholder(valor_visible_celda(ws, merged_lookup, row, col))
            ]
            if not placeholder_cols:
                continue

            first_placeholder_col = min(placeholder_cols)
            detail = construir_detalle_prestacion(
                ws,
                merged_lookup,
                row,
                first_placeholder_col,
                descriptor_context,
            )
            if not detail:
                detail = code

            code_to_detail[code] = detail
            codes.append(code)

            prestaciones_rows.append(
                {
                    "rem": rem,
                    "seccion": key,
                    "seccion_nombre": section_name,
                    "codigo_prestacion": code,
                    "detalle_prestacion": detail,
                    "detalle_niveles": detail,
                    "fila_excel": row,
                }
            )

            for placeholder, label in sorted(
                columns_by_placeholder.items(),
                key=lambda item: indice_placeholder(item[0]),
            ):
                col_idx = placeholder_to_col[placeholder]
                niveles = [parte for parte in label.split(" - ") if parte]
                columnas_rows.append(
                    {
                        "rem": rem,
                        "seccion": key,
                        "seccion_nombre": section_name,
                        "codigo_prestacion": code,
                        "detalle_prestacion": detail,
                        "codigo_columna": placeholder,
                        "columna_excel": get_column_letter(col_idx),
                        "columna_indice_excel": col_idx,
                        "columna_detalle": label,
                        "columna_categoria": niveles[0] if niveles else "",
                        "columna_subcategoria": niveles[1] if len(niveles) > 1 else "",
                        "columna_sub_subcategoria": niveles[2] if len(niveles) > 2 else "",
                        "columna_nivel_4": niveles[3] if len(niveles) > 3 else "",
                        "columna_nivel_5": niveles[4] if len(niveles) > 4 else "",
                        "columna_niveles": niveles,
                        "fila_excel": row,
                    }
                )

        plantillas.append(
            SectionTemplate(
                key=key,
                name=section_name,
                title=title,
                start_row=start_row,
                end_row=hard_end,
                first_code_row=first_code_row,
                last_code_row=last_code_row,
                row_to_code=row_to_code,
                code_to_detail=code_to_detail,
                codes=codes,
                placeholder_to_col=placeholder_to_col,
                columns_by_placeholder=columns_by_placeholder,
            )
        )

    return plantillas, prestaciones_rows, columnas_rows


def guardar_json(path: Path, data: object) -> None:
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def exportar_diccionario_completo(rem: str) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, str], dict[str, dict[str, str]], dict[str, Path]]:
    plantillas, prestaciones_rows, columnas_rows = extraer_plantilla_rem(rem, DICT_PATH, rem)
    paths = paths_metadata(rem)

    seccion_a_codigos = {template.key: template.codes for template in plantillas}
    codigo_a_detalle: dict[str, str] = {}
    codigo_a_seccion: dict[str, str] = {}
    columnas_por_seccion = {
        template.key: template.columns_by_placeholder for template in plantillas
    }
    columnas_por_codigo: dict[str, dict[str, str]] = {}
    columnas_completas_por_codigo: dict[str, list[dict]] = defaultdict(list)

    for template in plantillas:
        for code, detail in template.code_to_detail.items():
            codigo_a_detalle[code] = detail
            codigo_a_seccion[code] = template.key
            columnas_por_codigo[code] = template.columns_by_placeholder

    for row in columnas_rows:
        columnas_completas_por_codigo[row["codigo_prestacion"]].append(row)

    metadata_base = {
        "rem": rem,
        "anio": int(ANIO),
        "archivo_fuente": str(DICT_PATH),
        "hoja": rem,
        "generado_en": datetime.now().isoformat(timespec="seconds"),
    }

    columnas_json = {
        "metadata": metadata_base,
        "columnas_por_seccion": columnas_por_seccion,
        "columnas_por_codigo": columnas_por_codigo,
        "columnas_completas_por_codigo": dict(columnas_completas_por_codigo),
    }
    prestaciones_json = {
        "metadata": metadata_base,
        "secciones": {
            template.key: {
                "nombre": template.name,
                "titulo": template.title,
                "fila_inicio": template.start_row,
                "fila_fin": template.last_code_row,
                "codigos": template.codes,
                "columnas": template.columns_by_placeholder,
            }
            for template in plantillas
        },
        "seccion_a_codigos": seccion_a_codigos,
        "codigo_a_detalle": codigo_a_detalle,
        "codigo_a_seccion": codigo_a_seccion,
        "prestaciones_por_codigo": {
            row["codigo_prestacion"]: row for row in prestaciones_rows
        },
        "prestaciones": prestaciones_rows,
    }

    guardar_json(paths["secciones"], seccion_a_codigos)
    guardar_json(paths["detalle"], codigo_a_detalle)
    guardar_json(paths["columnas_json"], columnas_json)
    guardar_json(paths["prestaciones_json"], prestaciones_json)

    prestaciones_df = pd.DataFrame(prestaciones_rows)
    columnas_df = pd.DataFrame(columnas_rows)
    prestaciones_df.to_csv(paths["prestaciones_csv"], index=False, encoding=CSV_ENCODING)
    columnas_df.to_csv(paths["columnas_csv"], index=False, encoding=CSV_ENCODING)

    return prestaciones_df, columnas_df, codigo_a_detalle, columnas_por_codigo, paths


def exportar_subconjunto_cardiovascular(
    prestaciones_df: pd.DataFrame,
    columnas_df: pd.DataFrame,
    codigo_a_detalle: dict[str, str],
    columnas_por_codigo: dict[str, dict[str, str]],
) -> list[str]:
    codigos_requeridos = []
    vistos = set()
    for item in INDICADORES_REM_P4:
        codigo = item["codigo_prestacion"]
        if codigo not in vistos:
            vistos.add(codigo)
            codigos_requeridos.append(codigo)

    subset_prestaciones = prestaciones_df[
        prestaciones_df["codigo_prestacion"].isin(codigos_requeridos)
    ].copy()
    subset_columnas = columnas_df[
        columnas_df["codigo_prestacion"].isin(codigos_requeridos)
    ].copy()

    subset_prestaciones.to_csv(
        DICCIONARIO_DIR / "REM_P4_2025_cardiovascular_prestaciones.csv",
        index=False,
        encoding=CSV_ENCODING,
    )
    subset_columnas.to_csv(
        DICCIONARIO_DIR / "REM_P4_2025_cardiovascular_columnas_por_codigo.csv",
        index=False,
        encoding=CSV_ENCODING,
    )

    subset_json = {
        "metadata": {
            "rem": REM,
            "anio": int(ANIO),
            "archivo_fuente": str(DICT_PATH),
            "archivo_indicadores": str(INDICADORES_PATH),
            "generado_en": datetime.now().isoformat(timespec="seconds"),
        },
        "codigos_requeridos": codigos_requeridos,
        "codigo_a_detalle": {codigo: codigo_a_detalle[codigo] for codigo in codigos_requeridos},
        "columnas_por_codigo": {
            codigo: columnas_por_codigo[codigo] for codigo in codigos_requeridos
        },
    }
    guardar_json(
        DICCIONARIO_DIR / "REM_P4_2025_cardiovascular_diccionario_resumido.json",
        subset_json,
    )

    return codigos_requeridos


def exportar_mapeo_indicadores(
    prestaciones_df: pd.DataFrame,
    columnas_df: pd.DataFrame,
    codigos_requeridos: list[str],
) -> None:
    codigo_to_info = (
        prestaciones_df.drop_duplicates(subset=["codigo_prestacion"])
        .set_index("codigo_prestacion")
        .to_dict("index")
    )
    columnas_por_prestacion = (
        columnas_df.groupby("codigo_prestacion")["codigo_columna"].nunique().to_dict()
    )

    rem_rows = []
    for item in INDICADORES_REM_P4:
        codigo = item["codigo_prestacion"]
        if codigo not in codigo_to_info:
            raise KeyError(f"No se encontró el código {codigo} en el diccionario P4.")
        info = codigo_to_info[codigo]
        rem_rows.append(
            {
                "indicador": item["indicador"],
                "nombre_indicador": item["nombre"],
                "origen": "REM P4",
                "rol": item["rol"],
                "rem": info["rem"],
                "seccion": info["seccion"],
                "seccion_nombre": info["seccion_nombre"],
                "codigo_prestacion": codigo,
                "detalle_prestacion": info["detalle_prestacion"],
                "fila_diccionario_rem": info["fila_excel"],
                "columnas_disponibles": columnas_por_prestacion.get(codigo, 0),
                "nota_metodologica": item["nota_metodologica"],
            }
        )

    no_rem_rows = [
        {
            "indicador": item["indicador"],
            "nombre_indicador": item["nombre"],
            "origen": item["origen"],
            "rol": "",
            "rem": "",
            "seccion": "",
            "seccion_nombre": "",
            "codigo_prestacion": "",
            "detalle_prestacion": "",
            "fila_diccionario_rem": "",
            "columnas_disponibles": "",
            "nota_metodologica": item["fuente"],
        }
        for item in INDICADORES_NO_REM
    ]

    mapeo_df = pd.DataFrame(rem_rows + no_rem_rows).sort_values(
        by=["indicador", "origen", "rol", "codigo_prestacion"],
        kind="stable",
    )
    mapeo_df.to_csv(
        DICCIONARIO_DIR / "indicadores_cardiovascular_2026_origenes_y_codigos.csv",
        index=False,
        encoding=CSV_ENCODING,
    )

    resumen = {
        "metadata": {
            "generado_en": datetime.now().isoformat(timespec="seconds"),
            "archivo_indicadores": str(INDICADORES_PATH),
            "rem_principal": REM,
            "anio_diccionario": int(ANIO),
        },
        "indicadores_con_rem": sorted({item["indicador"] for item in INDICADORES_REM_P4}),
        "indicadores_sin_rem": [item["indicador"] for item in INDICADORES_NO_REM],
        "codigos_requeridos": codigos_requeridos,
        "detalle": mapeo_df.to_dict(orient="records"),
    }
    guardar_json(
        DICCIONARIO_DIR / "indicadores_cardiovascular_2026_origenes_y_codigos.json",
        resumen,
    )


def exportar_resumen_markdown(prestaciones_df: pd.DataFrame, columnas_df: pd.DataFrame, codigos_requeridos: list[str]) -> None:
    secciones = prestaciones_df.groupby("seccion")["codigo_prestacion"].nunique().to_dict()
    columnas_por_codigo = (
        columnas_df.groupby("codigo_prestacion")["codigo_columna"].nunique().to_dict()
    )
    indicadores_rem = sorted({item["indicador"] for item in INDICADORES_REM_P4})
    indicadores_no_rem = [item["indicador"] for item in INDICADORES_NO_REM]

    contenido = "\n".join(
        [
            "# Diccionario REM Cardiovascular",
            "",
            f"- Excel revisado: `{INDICADORES_PATH.name}`",
            f"- REM identificado para indicadores clínicos: `{REM}`",
            f"- Archivo oficial de diccionario usado: `{DICT_PATH.name}`",
            f"- Prestaciones extraídas en `{REM}`: `{prestaciones_df['codigo_prestacion'].nunique()}`",
            f"- Columnas por prestación en `{REM}`: `{columnas_df['codigo_columna'].nunique()}`",
            f"- Secciones detectadas: `{', '.join(f'{k} ({v} códigos)' for k, v in secciones.items())}`",
            f"- Indicadores que usan REM: `{', '.join(str(x) for x in indicadores_rem)}`",
            f"- Indicadores sin REM directo: `{', '.join(str(x) for x in indicadores_no_rem)}`",
            "",
            "## Subconjunto cardiovascular",
            "",
            f"- Códigos REM P4 requeridos por la planilla: `{len(codigos_requeridos)}`",
            "",
        ]
        + [
            f"- `{codigo}`: {prestaciones_df.loc[prestaciones_df['codigo_prestacion'] == codigo, 'detalle_prestacion'].iloc[0]} (`{columnas_por_codigo.get(codigo, 0)}` columnas)"
            for codigo in codigos_requeridos
        ]
    )
    (DICCIONARIO_DIR / "README_diccionario.md").write_text(contenido, encoding="utf-8")


def main() -> None:
    DICCIONARIO_DIR.mkdir(parents=True, exist_ok=True)

    prestaciones_df, columnas_df, codigo_a_detalle, columnas_por_codigo, _ = exportar_diccionario_completo(REM)
    codigos_requeridos = exportar_subconjunto_cardiovascular(
        prestaciones_df,
        columnas_df,
        codigo_a_detalle,
        columnas_por_codigo,
    )
    exportar_mapeo_indicadores(prestaciones_df, columnas_df, codigos_requeridos)
    exportar_resumen_markdown(prestaciones_df, columnas_df, codigos_requeridos)

    print(
        json.dumps(
            {
                "rem": REM,
                "anio": int(ANIO),
                "prestaciones_extraidas": int(prestaciones_df["codigo_prestacion"].nunique()),
                "filas_prestaciones": int(len(prestaciones_df)),
                "filas_columnas": int(len(columnas_df)),
                "codigos_cardiovascular": len(codigos_requeridos),
                "directorio_salida": str(DICCIONARIO_DIR),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
